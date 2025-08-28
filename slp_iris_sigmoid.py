#!/usr/bin/env python3
"""
SLP (Sigmoid + MSE Delta Rule) untuk Iris (Setosa vs Versicolor)

Mencerminkan metode Google Sheet:
- Inisialisasi: theta0..theta4 = 0.5 (bias & 4 bobot), eta default 0.1
- 5 epoch, 100 langkah/epoch, urutan tetap (tanpa shuffle)
- Dua mode metrik:
  * streaming : akurasi/MSE dihitung "sebelum update" di setiap langkah (persis Excel)
  * snapshot  : akurasi/MSE dihitung dengan parameter akhir epoch (umum di ML modern)

Input:
- --xlsx: file .xlsx berisi sheet "Data" (kolom: x1,x2,x3,x4,species)
- --csv : file .csv (boleh: x1..x4 + species | x1..x4 + y | X1..X4 + TARGET [+ X0])

Output:
- CSV metrik per epoch (+ bobot & bias akhir tiap epoch)

Contoh:
python slp_iris_sigmoid.py --csv "./iris_train.csv" --epochs 5 --eta 0.1 \
  --metric_mode streaming --out slp_results.csv
"""

import argparse
from pathlib import Path
import numpy as np
import pandas as pd

# -----------------------
# Utils
# -----------------------
def sigmoid(z: np.ndarray | float) -> np.ndarray | float:
    return 1.0 / (1.0 + np.exp(-z))

def load_data(xlsx_path: str | None = None, csv_path: str | None = None) -> pd.DataFrame:
    """
    Mengembalikan DataFrame dengan kolom: x1,x2,x3,x4,y
    Mapping label: Iris-versicolor -> 1.0, Iris-setosa -> 0.0
    """
    if xlsx_path:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, data_only=True)
        if "Data" not in wb.sheetnames:
            raise ValueError("Sheet 'Data' tidak ditemukan di XLSX.")
        ws = wb["Data"]
        rows = []
        for r in range(1, ws.max_row + 1):
            species = ws.cell(r, 5).value
            if species is None:
                break
            x1 = float(ws.cell(r, 1).value)
            x2 = float(ws.cell(r, 2).value)
            x3 = float(ws.cell(r, 3).value)
            x4 = float(ws.cell(r, 4).value)
            y  = 1.0 if str(species).strip() == "Iris-versicolor" else 0.0
            rows.append((x1, x2, x3, x4, y))
        if not rows:
            raise ValueError("Sheet 'Data' kosong / format tidak sesuai.")
        df = pd.DataFrame(rows, columns=["x1","x2","x3","x4","y"])
        return df

    # CSV path
    if not csv_path:
        raise ValueError("Harus memberi --xlsx atau --csv.")
    df = pd.read_csv(csv_path)

    # Normalisasi kolom
    cols_lower = {c: c.lower() for c in df.columns}
    df = df.rename(columns=cols_lower)

    # Abaikan X0 kalau ada
    for maybe_x0 in ["x0", "bias", "theta0", "θ0"]:
        if maybe_x0 in df.columns:
            df = df.drop(columns=[maybe_x0])

    # Rename fitur jika uppercase
    rename_map = {}
    for k in ["x1","x2","x3","x4"]:
        if k not in df.columns and k.upper() in df.columns:
            rename_map[k.upper()] = k
    if rename_map:
        df = df.rename(columns=rename_map)

    # Label: y / target / species
    if "y" not in df.columns:
        if "target" in df.columns:
            df = df.rename(columns={"target": "y"})
        elif "species" in df.columns:
            df["y"] = (df["species"].astype(str).str.strip() == "Iris-versicolor").astype(float)
        else:
            raise ValueError("CSV harus mengandung salah satu kolom: y, target, atau species.")

    # Pastikan fitur ada
    for k in ["x1","x2","x3","x4"]:
        if k not in df.columns:
            raise ValueError(f"Kolom fitur '{k}' tidak ditemukan pada CSV.")
    df = df[["x1","x2","x3","x4","y"]].astype(float)
    return df

# -----------------------
# Training
# -----------------------
def train_sigmoid_mse(
    X: np.ndarray,
    y: np.ndarray,
    epochs: int = 5,
    eta: float = 0.1,
    metric_mode: str = "snapshot",
    steps_per_epoch: int = 100,
):
    """
    Training Sigmoid + MSE (delta rule) persis template Excel.

    metric_mode:
      - "streaming": hitung akurasi/MSE per langkah SEBELUM update, lalu update (match Excel).
      - "snapshot" : update dulu 100 langkah, lalu hitung akurasi/MSE dengan bobot akhir epoch.

    Return:
      params  : list[(w (4,), b)] akhir setiap epoch
      metrics : list[(train_acc, train_mse)] sesuai metric_mode
    """
    assert metric_mode in ("snapshot", "streaming")
    # init 0.5
    w = np.array([0.5, 0.5, 0.5, 0.5], dtype=float)
    b = 0.5

    params = []
    metrics = []

    if metric_mode == "streaming":
        # Mirroring Excel: log pred sebelum update, lalu update
        for ep in range(epochs):
            acc_sum = 0.0
            mse_sum = 0.0
            for i in range(steps_per_epoch):
                xi = X[i]
                yi = y[i]
                z = float(np.dot(w, xi) + b)
                p = sigmoid(z)
                q = 1.0 if p >= 0.5 else 0.0
                acc_sum += (q == yi)
                mse_sum += (p - yi) ** 2

                grad = 2 * (p - yi) * p * (1 - p)
                b -= eta * grad
                w -= eta * grad * xi

            params.append((w.copy(), float(b)))
            metrics.append((acc_sum / steps_per_epoch, mse_sum / steps_per_epoch))
        return params, metrics

    # snapshot mode: evaluasi pakai bobot akhir epoch
    for ep in range(epochs):
        # lakukan update 100 langkah
        for i in range(steps_per_epoch):
            xi = X[i]
            yi = y[i]
            z = float(np.dot(w, xi) + b)
            p = sigmoid(z)
            grad = 2 * (p - yi) * p * (1 - p)
            b -= eta * grad
            w -= eta * grad * xi

        params.append((w.copy(), float(b)))

        # evaluasi train (snapshot): pakai bobot akhir epoch
        p_tr = sigmoid(X[:steps_per_epoch] @ w + b)
        q_tr = (p_tr >= 0.5).astype(float)
        acc_tr = (q_tr == y[:steps_per_epoch]).mean()
        mse_tr = np.mean((p_tr - y[:steps_per_epoch]) ** 2)
        metrics.append((acc_tr, mse_tr))

    return params, metrics

def eval_val(params, Xv, yv):
    vals = []
    for (w, b) in params:
        p = sigmoid(Xv @ w + b)
        q = (p >= 0.5).astype(float)
        acc = (q == yv).mean()
        mse = np.mean((p - yv) ** 2)
        vals.append((acc, mse))
    return vals

# -----------------------
# Main
# -----------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=str, default=None, help="Path .xlsx (sheet 'Data')")
    ap.add_argument("--csv", type=str, default=None, help="Path .csv (x1..x4 + species/target/y)")
    ap.add_argument("--epochs", type=int, default=5)
    ap.add_argument("--eta", type=float, default=0.1)
    ap.add_argument("--out", type=str, default="slp_sigmoid_results.csv")
    ap.add_argument("--metric_mode", type=str, choices=["snapshot","streaming"],
                    default="snapshot", help="Cara hitung metrik train")
    ap.add_argument("--steps_per_epoch", type=int, default=100, help="Jumlah langkah/epoch (default 100)")
    ap.add_argument("--val_start", type=int, default=80, help="Index awal validasi (default 80)")
    ap.add_argument("--val_count", type=int, default=20, help="Jumlah data validasi (default 20)")
    ap.add_argument("--print_params", action="store_true", help="Cetak w,b tiap epoch untuk cross-check")
    args = ap.parse_args()

    df = load_data(xlsx_path=args.xlsx, csv_path=args.csv)
    X = df[["x1","x2","x3","x4"]].to_numpy()
    y = df["y"].to_numpy()

    # Train
    params, train_metrics = train_sigmoid_mse(
        X, y,
        epochs=args.epochs,
        eta=args.eta,
        metric_mode=args.metric_mode,
        steps_per_epoch=args.steps_per_epoch
    )

    # Validation slice (default 80..99)
    vs, vc = args.val_start, args.val_count
    Xv = X[vs:vs+vc]
    yv = y[vs:vs+vc]
    val_metrics = eval_val(params, Xv, yv)

    # Output CSV
    out_df = pd.DataFrame({
        "epoch": list(range(1, args.epochs+1)),
        "train_acc": [a for (a,m) in train_metrics],
        "train_mse": [m for (a,m) in train_metrics],
        "val_acc":   [a for (a,m) in val_metrics],
        "val_mse":   [m for (a,m) in val_metrics],
        "w":         [p[0].tolist() for p in params],
        "b":         [p[1] for p in params],
    })
    out_path = Path(args.out)
    out_df.to_csv(out_path, index=False)

    # Optional print
    if args.print_params:
        for i, (w, b) in enumerate(params, start=1):
            print(f"[Epoch {i}] b={b:.8f}, w={w.tolist()}")

    print(f"Saved: {out_path.resolve()}")

if __name__ == "__main__":
    main()
